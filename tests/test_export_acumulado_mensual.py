from dataclasses import replace
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock

from openpyxl import load_workbook
from sqlalchemy import create_engine

from cobranzas.application.chain.export_acumulado_handler import ExportAcumuladoHandler
from cobranzas.application.chain.proceso_context import ProcesoContext
from cobranzas.domain.models.credito import Credito
from cobranzas.domain.models.fila_acumulado_mensual import FilaAcumuladoMensual
from cobranzas.domain.services.exportar_acumulado_mensual_service import (
    ExportarAcumuladoMensualService,
)
from cobranzas.infrastructure.adapters.excel_acumulado_writer import (
    COLUMNAS,
    ExcelAcumuladoWriter,
)
from cobranzas.infrastructure.config.entregables_mensuales import (
    ruta_acumulado_mensual,
    ruta_asignacion_desde_fecha_archivo,
    ruta_asignacion_mensual,
)
from cobranzas.infrastructure.persistence.base import Base
from cobranzas.infrastructure.persistence.mappers.cobranza_credito_mapper import (
    ESTADO_ASESOR_MORA_TEMPRANA,
)
from cobranzas.infrastructure.persistence.mappers.deuda_deudor_mapper import mapear_deuda
from cobranzas.infrastructure.persistence.models import (
    Asesor,
    AsesorDeuda,
    Catalogo,
    Clave,
    Deuda,
    Deudor,
)
from cobranzas.infrastructure.persistence.repositories.acumulado_mensual_repository import (
    SqlAlchemyAcumuladoMensualRepository,
)
from cobranzas.infrastructure.persistence.repositories.cobranza_repository import (
    SqlAlchemyCobranzaRepository,
)
from cobranzas.infrastructure.persistence.session import get_session_factory
from sqlalchemy import select


def _fila_ejemplo(fecha: date, operacion: str = "001") -> FilaAcumuladoMensual:
    return FilaAcumuladoMensual(
        fecha_proceso=fecha,
        desc_oficina="OFICINA CENTRAL",
        nombre="CLIENTE UNO",
        cedula="1234567890",
        operacion=operacion,
        tipo_oper="CONSUMO",
        fecha_concesion=date(2020, 1, 15),
        valor_ori_prestamo=Decimal("10000"),
        saldo_cap_prest=Decimal("5000"),
        calificac="A",
        total_provision=Decimal("100"),
        saldo_140x=Decimal("10"),
        saldo_141x=Decimal("20"),
        saldo_142x=Decimal("30"),
        total_op=Decimal("5500"),
        est="VIGENTE",
        estado_mora=ESTADO_ASESOR_MORA_TEMPRANA,
        dias_atraso_camorosico=3,
        dias_mora=1,
        tipo="T1",
        dia_pago=5,
        valor_cuota=Decimal("250"),
        cuota_act=12,
        dividendos=24,
        oficial_adm="OF ADM",
        decision="SI",
        segmentacion="SEG1",
        score="700",
        fuente_repago="NOMINA",
        identificacion_ifi="IFI-1",
        usuario_asesor="520",
        nombre_asesor="ASESOR A",
        id_credito_recblue="99887",
    )


def test_mapear_deuda_separa_dias_camorosico_y_mora():
    credito = Credito(
        id_credito="001",
        cliente="Cliente",
        saldo_pendiente=100.0,
        dias_mora=1,
        fecha_corte=date(2026, 5, 6),
        campos_tab=(("dias_atraso", "3"),),
    )
    datos = mapear_deuda(credito)
    assert datos.dias_mora == 1
    assert datos.dias_atraso_camorosico == 3


def test_ruta_asignacion_usa_fecha_proceso(tmp_path: Path):
    feriados: set[date] = set()
    assert ruta_asignacion_desde_fecha_archivo(
        tmp_path, date(2026, 6, 1), feriados
    ) == (tmp_path / "2026" / "06" / "ASIGNACION_06022026.csv")
    assert ruta_asignacion_desde_fecha_archivo(
        tmp_path, date(2026, 6, 5), feriados
    ) == (tmp_path / "2026" / "06" / "ASIGNACION_06082026.csv")


def test_rutas_entregables_mensuales(tmp_path: Path):
    fecha = date(2026, 5, 6)
    assert ruta_acumulado_mensual(tmp_path, fecha) == (
        tmp_path / "2026" / "05" / "asignacion_acumulado_202605.xlsx"
    )
    assert ruta_asignacion_mensual(tmp_path, fecha) == (
        tmp_path / "2026" / "05" / "ASIGNACION_05062026.csv"
    )


def test_excel_acumulado_una_fila_por_operacion(tmp_path: Path):
    archivo = tmp_path / "acumulado.xlsx"
    writer = ExcelAcumuladoWriter()
    dia1 = date(2026, 5, 5)
    dia2 = date(2026, 5, 6)

    writer.anexar_lote(archivo, dia1, [_fila_ejemplo(dia1, "001")])
    writer.anexar_lote(archivo, dia2, [_fila_ejemplo(dia2, "002")])
    writer.anexar_lote(archivo, dia1, [_fila_ejemplo(dia1, "003")])

    libro = load_workbook(archivo, read_only=True, data_only=True)
    hoja = libro.active
    filas = list(hoja.iter_rows(min_row=2, values_only=True))
    assert len(filas) == 3
    operaciones = {str(f[4]) for f in filas}
    assert operaciones == {"001", "002", "003"}
    libro.close()


def test_excel_acumulado_conserva_fecha_proceso_anterior(tmp_path: Path):
    """Una operación ya registrada en una fecha de proceso anterior (se conserva
    del corte previo) no debe moverse a la fecha de proceso más reciente."""
    archivo = tmp_path / "acumulado.xlsx"
    writer = ExcelAcumuladoWriter()
    dia1 = date(2026, 6, 1)
    dia2 = date(2026, 6, 2)

    writer.anexar_lote(archivo, dia1, [_fila_ejemplo(dia1, "001")])
    writer.anexar_lote(
        archivo,
        dia2,
        [
            replace(
                _fila_ejemplo(dia2, "001"),
                fecha_proceso=dia2,
                nombre="CLIENTE ACTUALIZADO",
            )
        ],
    )

    libro = load_workbook(archivo, read_only=True, data_only=True)
    filas = list(libro.active.iter_rows(min_row=2, values_only=True))
    libro.close()

    assert len(filas) == 1
    assert filas[0][2] == "CLIENTE UNO"
    assert _parsear_fecha_desde_excel(filas[0][0]) == dia1


def test_excel_acumulado_refresca_misma_fecha_proceso(tmp_path: Path):
    """Re-ejecutar el mismo corte (misma fecha de proceso) sí refresca los datos."""
    archivo = tmp_path / "acumulado.xlsx"
    writer = ExcelAcumuladoWriter()
    dia = date(2026, 6, 1)

    writer.anexar_lote(archivo, dia, [_fila_ejemplo(dia, "001")])
    writer.anexar_lote(
        archivo,
        dia,
        [replace(_fila_ejemplo(dia, "001"), nombre="CLIENTE ACTUALIZADO")],
    )

    libro = load_workbook(archivo, read_only=True, data_only=True)
    filas = list(libro.active.iter_rows(min_row=2, values_only=True))
    libro.close()

    assert len(filas) == 1
    assert filas[0][2] == "CLIENTE ACTUALIZADO"
    assert _parsear_fecha_desde_excel(filas[0][0]) == dia


def _parsear_fecha_desde_excel(valor) -> date:
    from datetime import datetime

    if isinstance(valor, datetime):
        return valor.date()
    return datetime.strptime(str(valor), "%m/%d/%Y").date()


def test_excel_acumulado_encabezados(tmp_path: Path):
    archivo = tmp_path / "acumulado.xlsx"
    ExcelAcumuladoWriter().anexar_lote(
        archivo, date(2026, 5, 5), [_fila_ejemplo(date(2026, 5, 5))]
    )
    libro = load_workbook(archivo, read_only=True, data_only=True)
    encabezados = [c.value for c in next(libro.active.iter_rows(max_row=1))]
    assert encabezados == list(COLUMNAS)
    libro.close()


def test_repositorio_acumulado_por_fecha_corte():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    session_factory = get_session_factory(engine)
    fecha = date(2026, 5, 6)

    with session_factory() as session:
        deudor = Deudor(documento="123", nombre="Cliente", socio="S1")
        session.add(deudor)
        session.flush()
        deuda = Deuda(
            id_deudor=deudor.id_deudor,
            numero_operacion="001",
            fecha_corte=fecha,
            desc_oficina="OF CENTRAL",
            nombre="Cliente",
            cedula="123",
            dias_mora=1,
            dias_atraso_camorosico=3,
        )
        session.add(deuda)
        session.flush()
        clave = Clave(clave="TEST", descripcion="test", vigente=True)
        session.add(clave)
        session.flush()
        catalogo = Catalogo(id_clave=clave.id_clave, valor="mora_temprana", vigencia=True)
        session.add(catalogo)
        session.flush()
        asesor = Asesor(nombre="Asesor A", cedula="OF-520", activo=True)
        session.add(asesor)
        session.flush()
        session.add(
            AsesorDeuda(
                id_catalogo=catalogo.id_catalogo,
                id_asesor=asesor.id_asesor,
                id_deuda=deuda.id_deuda,
                estado=ESTADO_ASESOR_MORA_TEMPRANA,
                id_credito_recblue="99887",
                fecha_asignacion=fecha,
            )
        )
        session.commit()

    repo = SqlAlchemyAcumuladoMensualRepository(session_factory)
    filas = repo.filas_por_fecha_corte(fecha)
    assert len(filas) == 1
    assert filas[0].operacion == "001"
    assert filas[0].estado_mora == ESTADO_ASESOR_MORA_TEMPRANA
    assert filas[0].dias_atraso_camorosico == 3
    assert filas[0].dias_mora == 1
    assert filas[0].usuario_asesor == "520"
    assert filas[0].id_credito_recblue == "99887"
    assert filas[0].fecha_proceso == fecha


def test_repositorio_acumulado_excluye_no_mora_temprana():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    session_factory = get_session_factory(engine)
    fecha = date(2026, 5, 6)

    with session_factory() as session:
        deudor = Deudor(documento="456", nombre="Cliente 2", socio="S2")
        session.add(deudor)
        session.flush()
        deuda = Deuda(
            id_deudor=deudor.id_deudor,
            numero_operacion="002",
            fecha_corte=fecha,
        )
        session.add(deuda)
        session.flush()
        clave = Clave(clave="TEST2", descripcion="test", vigente=True)
        session.add(clave)
        session.flush()
        catalogo = Catalogo(id_clave=clave.id_clave, valor="mora_leve", vigencia=True)
        session.add(catalogo)
        session.flush()
        asesor = Asesor(nombre="Asesor B", cedula="OF-521", activo=True)
        session.add(asesor)
        session.flush()
        session.add(
            AsesorDeuda(
                id_catalogo=catalogo.id_catalogo,
                id_asesor=asesor.id_asesor,
                id_deuda=deuda.id_deuda,
                estado="MORA_MADURA",
                fecha_asignacion=fecha,
            )
        )
        session.commit()

    filas = SqlAlchemyAcumuladoMensualRepository(session_factory).filas_por_fecha_corte(
        fecha
    )
    assert filas == []


def test_exportar_acumulado_fecha_proceso_es_consulta_efectiva(tmp_path: Path):
    """Archivo 2-jun → FECHA DEL PROCESO 3-jun; archivo viernes 5-jun → lunes 8-jun."""
    repo = MagicMock()
    writer = ExcelAcumuladoWriter()
    service = ExportarAcumuladoMensualService(repo, writer, tmp_path)
    feriados: set[date] = set()

    repo.filas_por_fecha_corte.return_value = [_fila_ejemplo(date(2026, 6, 2), "001")]
    archivo = service.exportar(date(2026, 6, 2), feriados)
    libro = load_workbook(archivo, read_only=True, data_only=True)
    fila = next(libro.active.iter_rows(min_row=2, values_only=True))
    libro.close()
    assert _parsear_fecha_desde_excel(fila[0]) == date(2026, 6, 3)

    repo.filas_por_fecha_corte.return_value = [_fila_ejemplo(date(2026, 6, 5), "002")]
    archivo_vie = service.exportar(date(2026, 6, 5), feriados)
    libro = load_workbook(archivo_vie, read_only=True, data_only=True)
    filas = list(libro.active.iter_rows(min_row=2, values_only=True))
    libro.close()
    por_op = {str(f[4]): _parsear_fecha_desde_excel(f[0]) for f in filas}
    assert por_op["001"] == date(2026, 6, 3)
    assert por_op["002"] == date(2026, 6, 8)


def test_export_acumulado_handler_solo_con_persistencia():
    export = MagicMock()
    export.exportar.return_value = Path("destino/2026/05/acumulado.xlsx")
    feriados = MagicMock()
    feriados.fechas_vigentes.return_value = set()
    handler = ExportAcumuladoHandler(export, feriados)
    ctx = ProcesoContext(dias_mora_minimo=1, usar_mora_temprana=True, persistir_en_bd=True)
    ctx.registros_persistidos_bd = 5
    ctx.creditos_mora = [
        Credito("001", "C", 100.0, 1, date(2026, 5, 6)),
    ]

    resultado = handler._procesar(ctx)

    export.exportar.assert_called_once_with(date(2026, 5, 6), set())
    assert resultado.archivo_acumulado_mensual == Path("destino/2026/05/acumulado.xlsx")


def test_persistir_misma_operacion_en_dos_cortes_crea_filas_separadas():
    """Regresión: la misma operación en cortes distintos debe generar una
    deuda por corte (antes se 'robaba' la fila del corte anterior)."""
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    session_factory = get_session_factory(engine)
    repo = SqlAlchemyCobranzaRepository(session_factory, usar_mora_temprana=True)

    corte_may = date(2026, 5, 29)
    corte_jun = date(2026, 6, 1)

    def _credito(fecha: date) -> Credito:
        return Credito(
            id_credito="0011244704",
            cliente="CLIENTE",
            saldo_pendiente=100.0,
            dias_mora=1,
            fecha_corte=fecha,
            codigo_oficial="520",
            estado_operacion="VIGENTE",
        )

    repo.guardar_creditos_mora([_credito(corte_may)])
    repo.guardar_creditos_mora([_credito(corte_jun)])

    with session_factory() as session:
        deudas = session.scalars(
            select(Deuda).where(Deuda.numero_operacion == "0011244704")
        ).all()
        cortes = sorted(d.fecha_corte for d in deudas)
    assert cortes == [corte_may, corte_jun]

    acumulado = SqlAlchemyAcumuladoMensualRepository(session_factory)
    assert len(acumulado.filas_por_fecha_corte(corte_may)) == 1
    assert len(acumulado.filas_por_fecha_corte(corte_jun)) == 1


def test_exportar_acumulado_filtra_dias_mora_uno(tmp_path: Path):
    """Solo se almacenan operaciones con días de mora > 1 (>= 2 por default)."""
    repo = MagicMock()
    writer = ExcelAcumuladoWriter()
    service = ExportarAcumuladoMensualService(repo, writer, tmp_path)
    fecha = date(2026, 6, 2)

    repo.filas_por_fecha_corte.return_value = [
        replace(_fila_ejemplo(fecha, "001"), dias_atraso_camorosico=3, dias_mora=3),
        replace(_fila_ejemplo(fecha, "002"), dias_atraso_camorosico=1, dias_mora=1),
        replace(_fila_ejemplo(fecha, "003"), dias_atraso_camorosico=2, dias_mora=0),
    ]
    archivo = service.exportar(fecha, set())

    libro = load_workbook(archivo, read_only=True, data_only=True)
    filas = list(libro.active.iter_rows(min_row=2, values_only=True))
    libro.close()
    operaciones = {str(f[4]) for f in filas}
    assert operaciones == {"001", "003"}


def test_exportar_acumulado_usa_dias_mora_si_no_hay_camorosico(tmp_path: Path):
    """Si no hay días CAMOROSICO, se evalúa dias_mora."""
    repo = MagicMock()
    writer = ExcelAcumuladoWriter()
    service = ExportarAcumuladoMensualService(repo, writer, tmp_path)
    fecha = date(2026, 6, 2)

    repo.filas_por_fecha_corte.return_value = [
        replace(_fila_ejemplo(fecha, "001"), dias_atraso_camorosico=None, dias_mora=5),
        replace(_fila_ejemplo(fecha, "002"), dias_atraso_camorosico=None, dias_mora=1),
    ]
    archivo = service.exportar(fecha, set())

    libro = load_workbook(archivo, read_only=True, data_only=True)
    filas = list(libro.active.iter_rows(min_row=2, values_only=True))
    libro.close()
    operaciones = {str(f[4]) for f in filas}
    assert operaciones == {"001"}


def test_exportar_acumulado_omite_si_todas_fuera_de_rango(tmp_path: Path):
    repo = MagicMock()
    writer = ExcelAcumuladoWriter()
    service = ExportarAcumuladoMensualService(repo, writer, tmp_path)
    fecha = date(2026, 6, 2)

    repo.filas_por_fecha_corte.return_value = [
        replace(_fila_ejemplo(fecha, "001"), dias_atraso_camorosico=1, dias_mora=1),
    ]
    archivo = service.exportar(fecha, set())

    assert not archivo.exists()


def test_export_acumulado_handler_omite_sin_persistencia():
    export = MagicMock()
    handler = ExportAcumuladoHandler(export)
    ctx = ProcesoContext(dias_mora_minimo=1, usar_mora_temprana=True, persistir_en_bd=False)

    handler._procesar(ctx)

    export.exportar.assert_not_called()
