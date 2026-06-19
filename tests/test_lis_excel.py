from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from cobranzas.infrastructure.adapters.lis_excel_writer import LisExcelWriter
from cobranzas.jobs.lis_excel_runner import ejecutar_lis_a_excel


def _escribir_lis(ruta: Path) -> None:
    lineas = [
        "CUADRO DE MOROSIDAD - CONSOLIDADO",
        "CORTE A: 06/01/2026",
        "OFICINA\tNO.OPERACION\tDIAS ATRASO",
        "1\t0012117441\t8",
        "1\t0015219214\t134",
    ]
    ruta.write_text("\n".join(lineas), encoding="utf-8")


def test_convertir_preserva_lineas_y_ceros_izquierda(tmp_path: Path):
    origen = tmp_path / "camorosico_06012026_2324_of_0.lis"
    _escribir_lis(origen)
    destino = tmp_path / "out" / "camorosico.xlsx"

    filas = LisExcelWriter().convertir(origen, destino)

    assert filas == 5
    assert destino.is_file()
    hoja = load_workbook(destino, read_only=True).active
    valores = [tuple(f) for f in hoja.iter_rows(values_only=True)]
    assert valores[2] == ("OFICINA", "NO.OPERACION", "DIAS ATRASO")
    # ceros a la izquierda preservados (texto)
    assert valores[3][1] == "0012117441"
    assert valores[3][2] == "8"


def test_runner_convierte_camorosico_y_cadetacaco(tmp_path: Path):
    cam = tmp_path / "lote" / "camorosico_06012026.lis"
    cad = tmp_path / "lote" / "cadetacaco_cie06012026.lis"
    cam.parent.mkdir(parents=True, exist_ok=True)
    _escribir_lis(cam)
    _escribir_lis(cad)
    destino = tmp_path / "destino"

    cfg = SimpleNamespace(
        archivo_morosidad=cam,
        archivo_cartera=cad,
        directorio_destino=destino,
        fecha_corte="06012026",
        log_level="INFO",
    )

    resultado = ejecutar_lis_a_excel(settings=cfg, configurar_logs=False)

    assert resultado.ok
    assert resultado.codigo_salida == 0
    assert len(resultado.archivos) == 2
    destinos = {Path(a.destino).name for a in resultado.archivos}
    assert destinos == {"camorosico_06012026.xlsx", "cadetacaco_cie06012026.xlsx"}
    for a in resultado.archivos:
        assert Path(a.destino).is_file()
        assert Path(a.destino).parent == destino / "excel_lis"
        assert a.filas == 5


def test_runner_sin_archivos_devuelve_error(tmp_path: Path):
    cfg = SimpleNamespace(
        archivo_morosidad=tmp_path / "no_existe_cam.lis",
        archivo_cartera=tmp_path / "no_existe_cad.lis",
        directorio_destino=tmp_path / "destino",
        fecha_corte="06012026",
        log_level="INFO",
    )

    resultado = ejecutar_lis_a_excel(settings=cfg, configurar_logs=False)

    assert not resultado.ok
    assert resultado.codigo_salida == 1
    assert resultado.archivos == ()
