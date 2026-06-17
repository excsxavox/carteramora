from datetime import date
from pathlib import Path
from unittest.mock import MagicMock

from openpyxl import load_workbook

from cobranzas.application.chain.export_acumulado_fin_mes_handler import (
    ExportAcumuladoFinMesHandler,
)
from cobranzas.application.chain.proceso_context import ProcesoContext
from cobranzas.domain.models.credito import Credito
from cobranzas.domain.services.exportar_acumulado_fin_mes_service import (
    ExportarAcumuladoFinMesService,
)
from cobranzas.infrastructure.adapters.excel_acumulado_fin_mes_writer import (
    ENCABEZADOS,
    ExcelAcumuladoFinMesWriter,
)
from cobranzas.infrastructure.config.entregables_mensuales import ruta_acumulado_fin_mes
from cobranzas.infrastructure.persistence.mappers.acumulado_fin_mes_mapper import (
    credito_a_fila_acumulado_fin_mes,
)


def _credito_mora() -> Credito:
    return Credito(
        id_credito="0016796922",
        cliente="CLIENTE FIN MES",
        saldo_pendiente=500.0,
        dias_mora=45,
        fecha_corte=date(2026, 6, 4),
        cedula="1234567890",
        socio="77370",
        calificacion="A",
        total_operacion=10000.0,
        campos_tab=(
            ("desc_oficina", "OFICINA CENTRAL"),
            ("tipo_oper", "CONSUMO"),
            ("dias_atraso", "45"),
            ("est", "VIGENTE"),
            ("sector", "URBANO"),
            ("interes_mora", "12.50"),
            ("actividad_economica", "COMERCIO"),
        ),
    )


def test_ruta_acumulado_fin_mes(tmp_path: Path):
    fecha_proceso = date(2026, 6, 5)
    assert ruta_acumulado_fin_mes(tmp_path, fecha_proceso) == (
        tmp_path / "2026" / "06" / "acumulado_fin_mes_06052026.xlsx"
    )


def test_mapper_fin_mes_campos_deuda_deudores():
    fila = credito_a_fila_acumulado_fin_mes(
        _credito_mora(),
        date(2026, 6, 5),
        dias_mora_minimo=30,
        archivo_origen="docsmora/camorosico.lis",
    )
    assert fila.numero_operacion == "0016796922"
    assert fila.fecha_proceso == date(2026, 6, 5)
    assert fila.deudores_documento == "1234567890"
    assert fila.deudores_nombre == "CLIENTE FIN MES"
    assert fila.deudores_socio == "77370"
    assert fila.estado_mora == "MORA_LEVE"
    assert fila.dias_mora == 45
    assert fila.sector == "URBANO"
    assert fila.interes_mora is not None
    assert fila.actividad_economica == "COMERCIO"
    assert fila.archivo_origen == "docsmora/camorosico.lis"


def test_exportar_acumulado_fin_mes_fecha_proceso(tmp_path: Path):
    service = ExportarAcumuladoFinMesService(
        ExcelAcumuladoFinMesWriter(), tmp_path, dias_mora_minimo=30
    )
    archivo = service.exportar(
        [_credito_mora()],
        date(2026, 6, 4),
        set(),
    )
    assert archivo.name == "acumulado_fin_mes_06052026.xlsx"
    assert archivo.is_file()


def test_excel_fin_mes_encabezados_deuda_deudores(tmp_path: Path):
    archivo = tmp_path / "fin_mes.xlsx"
    ExcelAcumuladoFinMesWriter().anexar_lote(
        archivo,
        date(2026, 6, 4),
        [credito_a_fila_acumulado_fin_mes(_credito_mora(), date(2026, 6, 5), 30)],
    )
    libro = load_workbook(archivo, read_only=True, data_only=True)
    encabezados = [c.value for c in next(libro.active.iter_rows(max_row=1))]
    libro.close()
    assert encabezados == list(ENCABEZADOS)
    assert "DEUDORES_DOCUMENTO" in encabezados
    assert "INTERES_MORA" in encabezados
    assert "ACTIVIDAD_ECONOMICA" in encabezados
    assert "PROCESO_COD" in encabezados


def test_export_fin_mes_handler():
    export = MagicMock()
    export.exportar.return_value = Path("destino/2026/06/acumulado_fin_mes_06052026.xlsx")
    handler = ExportAcumuladoFinMesHandler(export, {date(2026, 1, 1)})
    ctx = ProcesoContext(dias_mora_minimo=30)
    ctx.archivo_morosidad = Path("docsmora/camorosico.lis")
    ctx.creditos_mora = [_credito_mora()]

    resultado = handler._procesar(ctx)

    export.exportar.assert_called_once()
    args, kwargs = export.exportar.call_args
    assert kwargs["archivo_origen"].replace("\\", "/") == "docsmora/camorosico.lis"
    assert resultado.archivo_acumulado_fin_mes == Path(
        "destino/2026/06/acumulado_fin_mes_06052026.xlsx"
    )


def test_excel_fin_mes_acumula_operaciones_mismo_dia(tmp_path: Path):
    writer = ExcelAcumuladoFinMesWriter()
    service = ExportarAcumuladoFinMesService(writer, tmp_path, dias_mora_minimo=30)
    credito1 = _credito_mora()
    credito2 = Credito(
        id_credito="0015219214",
        cliente="OTRO",
        saldo_pendiente=200.0,
        dias_mora=100,
        fecha_corte=date(2026, 6, 4),
        campos_tab=(("dias_atraso", "100"),),
    )

    service.exportar([credito1], date(2026, 6, 4), set())
    service.exportar([credito2], date(2026, 6, 4), set())

    archivo = tmp_path / "2026" / "06" / "acumulado_fin_mes_06052026.xlsx"
    libro = load_workbook(archivo, read_only=True, data_only=True)
    filas = list(libro.active.iter_rows(min_row=2, values_only=True))
    libro.close()
    idx_op = ENCABEZADOS.index("NUMERO_OPERACION")
    operaciones = {str(f[idx_op]) for f in filas}
    assert operaciones == {"0016796922", "0015219214"}
